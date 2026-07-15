#!/usr/bin/env python3
"""build_workbook.py — turn one or more match_statement.py proposal JSONs into a
single reviewable Excel workbook (Cover + one sheet per account).

This is a formatting/reporting step only — it makes no Xero calls and changes
nothing. Run match_statement.py first for each account, then this script.

Usage:
  python3 build_workbook.py --out 3CP_Bank_Reconciliation_<period>.xlsx \
    --account "DBS USD (000331663)" proposal_dbs_usd.json USD \
    --account "DBS HKD (000331627)" proposal_dbs_hkd.json HKD \
    --account "CCB USD (000013996642)" proposal_ccb_usd.json USD \
    --account "CCB HKD (000013996634)" proposal_ccb_hkd.json HKD \
    --period "01-Jun-2026 to 30-Jun-2026" \
    --out-of-scope "IBKR — broker NAV/cash reconciliation is a different workflow (ibkr_nav_check.py in 3cp-investment-treasury-close), not this bank-transaction matcher."

--account takes exactly 3 values: <sheet label> <proposal.json path> <currency>.
Repeat --account for each account you have a proposal for. An account with NO
posted transactions this period still gets a sheet — it will show 100%
unmatched statement lines, which is a correct, informative result (nothing to
reconcile against yet), not an error.

For an account you have NO proposal for at all (never ran the matcher — e.g.
no statement was available, or the account is genuinely out of scope, like a
broker NAV reconciliation), use --out-of-scope "<free text>" instead; it goes
on the Cover sheet as a stated exclusion rather than being silently omitted.
"""

import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E5F")
SUB_FONT = Font(name=FONT, size=10, italic=True, color="555555")
NORMAL = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
HIGH_FILL = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
MED_FILL = PatternFill("solid", start_color="FFE699", end_color="FFE699")
FEE_FILL = PatternFill("solid", start_color="FFD966", end_color="FFD966")
UNM_FILL = PatternFill("solid", start_color="F4B6B6", end_color="F4B6B6")
EMPTY_FILL = PatternFill("solid", start_color="E7E6E6", end_color="E7E6E6")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CUR = '#,##0.00;(#,##0.00);"-"'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        if w:
            ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, vals, fill=None, bold=False):
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = BOLD if bold else NORMAL
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if isinstance(v, (int, float)):
            cell.number_format = CUR
            cell.alignment = Alignment(horizontal="right")


def load_proposal(path):
    with open(path) as f:
        return json.load(f)


def build_account_sheet(wb, sheet_name, proposal_path, currency):
    p = load_proposal(proposal_path)
    ws = wb.create_sheet(sheet_name[:31])  # Excel sheet-name length limit
    ws["A1"] = f"{sheet_name} — {currency} — Match Proposal"
    ws["A1"].font = Font(name=FONT, bold=True, size=13, color="1F4E5F")
    ws.merge_cells("A1:L1")

    headers = ["Status", "Confidence", "Date (Stmt)", "Date (Txn)", "Date gap (d)",
               "Amount (Stmt)", "Amount (Txn)", "Payee / Contact", "Reference (Stmt)",
               "Reference (Txn)", "Xero Txn ID", "Reason / Note"]
    hr = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))
    ws.freeze_panes = f"A{hr + 1}"

    row = hr + 1
    any_rows = False

    for m in p.get("matched", []):
        any_rows = True
        sl, tx = m["statement_line"], m["transaction"]
        write_row(ws, row, ["MATCHED", m["confidence"].upper(), sl.get("date"), tx.get("date"),
                             m.get("date_gap_days") or "", sl.get("amount"), tx.get("amount"),
                             sl.get("payee") or tx.get("contact"), sl.get("reference"), tx.get("reference"),
                             tx.get("id"), m.get("reason")],
                  fill=HIGH_FILL if m["confidence"] == "high" else MED_FILL)
        row += 1

    for g in p.get("grouped", []):
        any_rows = True
        if "transactions" in g:
            sl = g["statement_line"]
            for tx in g["transactions"]:
                write_row(ws, row, ["GROUPED (1 stmt : many txn)", "-", sl.get("date"), tx.get("date"), "",
                                     sl.get("amount"), tx.get("amount"), sl.get("payee") or tx.get("contact"),
                                     sl.get("reference"), tx.get("reference"), tx.get("id"), g.get("reason")],
                          fill=MED_FILL)
                row += 1
        else:
            tx = g["transaction"]
            for sl in g["statement_lines"]:
                write_row(ws, row, ["GROUPED (many stmt : 1 txn)", "-", sl.get("date"), tx.get("date"), "",
                                     sl.get("amount"), tx.get("amount"), sl.get("payee") or tx.get("contact"),
                                     sl.get("reference"), tx.get("reference"), tx.get("id"), g.get("reason")],
                          fill=MED_FILL)
                row += 1

    for a in p.get("ambiguous", []):
        any_rows = True
        sl = a["statement_line"]
        for cand in a["candidates"]:
            write_row(ws, row, ["AMBIGUOUS — pick in Xero", "-", sl.get("date"), cand.get("date"), "",
                                 sl.get("amount"), cand.get("amount"), sl.get("payee") or cand.get("contact"),
                                 sl.get("reference"), cand.get("reference"), cand.get("id"), a.get("reason")],
                      fill=MED_FILL)
            row += 1

    for fl in p.get("possible_fee_flags", []):
        any_rows = True
        sl, tx = fl["statement_line"], fl["transaction"]
        write_row(ws, row, ["POSSIBLE FEE — verify", "-", sl.get("date"), tx.get("date"), "",
                             sl.get("amount"), tx.get("amount"), sl.get("payee") or tx.get("contact"),
                             sl.get("reference"), tx.get("reference"), tx.get("id"),
                             f"{fl.get('reason')} (diff {fl.get('amount_difference')})"],
                  fill=FEE_FILL)
        row += 1

    for sl in p.get("unmatched_statement_lines", []):
        any_rows = True
        write_row(ws, row, ["UNMATCHED (statement line)", "-", sl.get("date"), "", "",
                             sl.get("amount"), "", sl.get("payee"), sl.get("reference"), "", "",
                             "No posted transaction found — either not yet posted, or out of scope for this batch."],
                  fill=UNM_FILL)
        row += 1

    for tx in p.get("unmatched_transactions", []):
        any_rows = True
        write_row(ws, row, ["UNMATCHED (transaction)", "-", "", tx.get("date"), "",
                             "", tx.get("amount"), tx.get("contact"), "", tx.get("reference"), tx.get("id"),
                             "No statement line found — check timing (may settle next period) or verify the posting."],
                  fill=UNM_FILL)
        row += 1

    if not any_rows:
        write_row(ws, row, ["No statement lines and no open transactions for this account/period.",
                             "-", "", "", "", "", "", "", "", "", "", ""], fill=EMPTY_FILL)
        row += 1

    ws.auto_filter.ref = f"A{hr}:L{row - 1}"
    autosize(ws, [26, 12, 12, 12, 11, 14, 14, 24, 20, 20, 12, 46])

    s = p.get("summary", {})
    return {
        "sheet": sheet_name,
        "stmt_lines": s.get("statement_lines", 0),
        "open_txns": s.get("open_transactions", 0),
        "matched": s.get("matched_1to1", 0) + s.get("grouped", 0),
        "fee_flags": s.get("possible_fee_flags", 0),
        "unmatched_lines": s.get("unmatched_statement_lines", 0),
        "unmatched_txns": s.get("unmatched_transactions", 0),
    }


def main():
    ap = argparse.ArgumentParser(description="Build a reconciliation workbook from match_statement.py proposals.")
    ap.add_argument("--account", nargs=3, action="append", default=[],
                     metavar=("LABEL", "PROPOSAL_JSON", "CURRENCY"),
                     help="Repeatable. One sheet per account.")
    ap.add_argument("--out-of-scope", action="append", default=[],
                     help="Repeatable. Free-text note for an account/workflow deliberately not covered this run.")
    ap.add_argument("--period", default="", help="Period label for the Cover sheet.")
    ap.add_argument("--out", required=True, help="Output .xlsx path.")
    args = ap.parse_args()

    if not args.account:
        raise SystemExit("Provide at least one --account LABEL PROPOSAL_JSON CURRENCY.")

    FONT_ = FONT
    wb = Workbook()
    cov = wb.active
    cov.title = "Cover"
    cov["A1"] = "3 Capital Partners Limited"
    cov["A1"].font = TITLE_FONT
    cov["A2"] = "Bank Reconciliation Match Proposal"
    cov["A2"].font = Font(name=FONT_, bold=True, size=12)
    cov["A3"] = ("Generated by 3cp-bank-reconciliation-helper — READ-ONLY proposal. "
                 "Confirm each match in Xero's Reconcile tab; this workbook makes no changes to Xero.")
    cov["A3"].font = SUB_FONT
    cov["A3"].alignment = Alignment(wrap_text=True)
    cov.merge_cells("A3:F3")
    cov.row_dimensions[3].height = 30

    info_rows = [
        ("Period", args.period or "(not specified)"),
        ("Matching basis", "Signed amount + reference overlap, within one currency. NEVER matched on date — "
                            "date gap shown for information only."),
        ("Confidence tiers", "HIGH = amount + reference confirm; MEDIUM = amount uniquely matches but no reference "
                              "to confirm (verify payee/date); POSSIBLE FEE = amounts differ by a small amount "
                              "consistent with a wire/bank fee; UNMATCHED = no counterpart found."),
    ]
    r = 5
    for label, val in info_rows:
        cov.cell(row=r, column=1, value=label).font = BOLD
        cov.cell(row=r, column=2, value=val).font = NORMAL
        cov.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        cov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cov.row_dimensions[r].height = 30
        r += 1

    if args.out_of_scope:
        cov.cell(row=r, column=1, value="Out of scope this run").font = BOLD
        for note in args.out_of_scope:
            cov.cell(row=r, column=2, value=note).font = NORMAL
            cov.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            cov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            cov.row_dimensions[r].height = 26
            r += 1

    r += 1
    cov.cell(row=r, column=1, value="Summary").font = Font(name=FONT_, bold=True, size=12)
    r += 1
    summary_headers = ["Account", "Statement lines", "Open transactions", "Matched",
                        "Possible fee", "Unmatched lines", "Unmatched txns"]
    for c, h in enumerate(summary_headers, start=1):
        cov.cell(row=r, column=c, value=h)
    style_header(cov, r, len(summary_headers))
    r += 1

    for label, proposal_path, currency in args.account:
        stats = build_account_sheet(wb, label, proposal_path, currency)
        write_row(cov, r, [stats["sheet"], stats["stmt_lines"], stats["open_txns"], stats["matched"],
                            stats["fee_flags"], stats["unmatched_lines"], stats["unmatched_txns"]])
        r += 1

    autosize(cov, [26, 16, 16, 10, 12, 14, 14])
    wb.save(args.out)
    print(f"Saved: {args.out}")


if __name__ == "__main__":
    main()
