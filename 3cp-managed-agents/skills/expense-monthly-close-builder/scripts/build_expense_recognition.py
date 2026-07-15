#!/usr/bin/env python3
"""Build 3CP expense recognition for a month from the Expensify expense-level export.
Recognition at SUBMISSION: DR expense (GL Category) / CR 135 (Company Card) or 801 (Cash).
PE Program (Tag2) lines reclass DR -> 664. Status filter: Report Display Status == 'Outstanding'.
All amounts HKD (Report Currency). Outputs: ManualJournal Xero CSV, contract payload JSON,
review workpaper xlsx (summary, by-card, by-report open-item schedule, line detail, checklist).
"""
import csv, json, datetime, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC='/mnt/user-data/uploads/May_2026_Expensify.csv'
PERIOD='2026-05'; PEND='2026-05-31'; LBL='May 2026'
AMEX_CLEARING='135'; UNPAID_CLAIMS='801'; PEP_RECV='664'
OUT='/home/claude/build/out'; os.makedirs(OUT, exist_ok=True)

rows=list(csv.reader(open(SRC))); H=rows[0]; D=rows[1:]
def I(n): return H.index(n)
c_amt,c_rc,c_tt,c_rim,c_bo,c_tag2,c_card,c_rid,c_gl,c_cat,c_date,c_mer,c_sub,c_subdate,c_rstatus = \
 I('Amount'),I('Report Currency'),I('Transaction Type'),I('Reimbursable'),I('Report Display Status'),I('Tag2'),I('Card Details'),I('Report ID'),I('GL Category'),I('Category'),I('Date'),I('Merchant'),I('Submitter Name'),I('Report Submitted Date'),I('Report Status')
def num(x):
    try: return round(float(str(x).replace(',','')),2)
    except: return 0.0

lines=[]
for r in D:
    if r[c_bo].strip()!='Outstanding': continue          # exclude Draft
    pe = (r[c_tag2].strip()=='PE Program')
    dr_acct = PEP_RECV if pe else r[c_gl].strip()
    track = 'CARD' if r[c_tt].strip()=='Company Card' else 'CASH'
    lines.append(dict(amt=num(r[c_amt]), dr=dr_acct, track=track, pe=pe, card=r[c_card].strip(),
                      rid=r[c_rid].strip(), gl=r[c_gl].strip(), cat=r[c_cat].strip(),
                      date=r[c_date].strip(), mer=r[c_mer].strip(), sub=r[c_sub].strip(),
                      subdate=r[c_subdate].strip(), rstatus=r[c_rstatus].strip()))

card_t=round(sum(l['amt'] for l in lines if l['track']=='CARD'),2)
cash_t=round(sum(l['amt'] for l in lines if l['track']=='CASH'),2)
pep_t =round(sum(l['amt'] for l in lines if l['pe']),2)

# ---- journals (contract payload) ----
def grp(ls):
    g=defaultdict(float)
    for l in ls: g[l['dr']]+=l['amt']
    return {k:round(v,2) for k,v in g.items()}

journals=[]
# AMEX recognition (one journal): DR expense codes / CR 135
card=grp([l for l in lines if l['track']=='CARD'])
jl=[{'account':k,'account_code':k,'debit':v,'credit':0.0} for k,v in sorted(card.items())]
jl.append({'account':'AMEX Clearing Account','account_code':AMEX_CLEARING,'debit':0.0,'credit':round(sum(card.values()),2)})
journals.append(dict(reference=f'3CP-EXP-{PERIOD}-AMEX-RECOG',status='DRAFT',date=PEND,currency='HKD',
    narration=f'{LBL} AMEX card expenses recognised at submission (Expensify, Outstanding)',
    reversal={'flag':False,'date':None},lines=jl))
# Reimbursable recognition: one journal per Report ID -> CR 801
byrep=defaultdict(list)
for l in lines:
    if l['track']=='CASH': byrep[l['rid']].append(l)
for rid in sorted(byrep):
    g=grp(byrep[rid]); tot=round(sum(g.values()),2); sub=byrep[rid][0]['sub']
    jl=[{'account':k,'account_code':k,'debit':v,'credit':0.0} for k,v in sorted(g.items())]
    jl.append({'account':'Unpaid Expense Claims','account_code':UNPAID_CLAIMS,'debit':0.0,'credit':tot})
    journals.append(dict(reference=f'3CP-EXP-{PERIOD}-REIMB-{rid}',status='DRAFT',date=PEND,currency='HKD',
        narration=f'{LBL} reimbursable claim recognised at submission — {sub} ({rid})',
        reversal={'flag':False,'date':None},lines=jl))

payload=dict(contract_version='1.0',entity='3 Capital Partners Limited',source_skill='expense-monthly-close-builder',
    period=PERIOD,period_end=PEND,basis='recognition at submission (Outstanding); HKD',
    target={'system':'Xero','post_as':'DRAFT','auto_approve':False},journals=journals)
json.dump(payload,open(f'{OUT}/3CP_EXP_{PERIOD}_draft_journals.json','w'),indent=2)

# ---- ManualJournal Xero CSV ----
HDR=['*Narration','*Date','Description','*AccountCode','*TaxRate','*Amount','TrackingName1','TrackingOption1','TrackingName2','TrackingOption2']
d=datetime.date.fromisoformat(PEND).strftime('%d/%m/%Y')
csvrows=[]
for j in journals:
    narr=f"{j['reference']} | {j['narration'][:150]}"
    for ln in j['lines']:
        amt=ln['debit']-ln['credit']
        csvrows.append({'*Narration':narr,'*Date':d,'Description':ln['account'],'*AccountCode':ln['account_code'],
                        '*TaxRate':'Tax Exempt','*Amount':f'{amt:.2f}','TrackingName1':'','TrackingOption1':'','TrackingName2':'','TrackingOption2':''})
with open(f'{OUT}/3CP_EXP_{LBL.replace(" ","")}_ManualJournal_Xero.csv','w',newline='') as f:
    w=csv.DictWriter(f,fieldnames=HDR); w.writeheader()
    for r in csvrows: w.writerow(r)

print(f"Outstanding lines: {len(lines)}  | Card -> 135: {card_t:,.2f}  | Cash -> 801: {cash_t:,.2f}  | PE -> 664: {pep_t:,.2f}")
print(f"Journals: 1 AMEX + {len(byrep)} reimbursable (per report) = {len(journals)}")
print(f"MJ CSV lines: {len(csvrows)}")

# ================= review workpaper =================
NAVY='1F3864'; F=lambda **k:Font(name='Calibri',**k); HKD='#,##0.00;(#,##0.00);"-"'
sub=PatternFill('solid',fgColor='D9E1F2'); gfill=PatternFill('solid',fgColor='F2F2F2'); navy=PatternFill('solid',fgColor=NAVY)
thin=Side(style='thin',color='BFBFBF'); box=Border(left=thin,right=thin,top=thin,bottom=thin)
def sh(c): c.font=F(bold=True,color='FFFFFF'); c.fill=navy; c.alignment=Alignment(horizontal='center')
wb=Workbook(); wb.remove(wb.active)
# Cover
cv=wb.create_sheet('Cover'); cv.column_dimensions['A'].width=3; cv.column_dimensions['B'].width=26; cv.column_dimensions['C'].width=90
cv['B2']='3 Capital Partners Limited'; cv['B2'].font=F(bold=True,size=16)
cv['B3']=f'Expense close — {LBL} (recognition at submission)'; cv['B3'].font=F(bold=True,size=13,color=NAVY)
cv['B4']='HKD | DRAFT for review'; cv['B4'].font=F(italic=True); r=6
def row(k,v):
    global r; cv.cell(r,2,k).font=F(bold=True); cc=cv.cell(r,3,v); cc.alignment=Alignment(wrap_text=True,vertical='top'); r+=1
row('Source',f'Expensify expense-level export ({len(D)} rows; {len(lines)} Outstanding, Draft excluded). P&L account per line = GL Category (col K). All HKD (Report Currency).')
row('Recognition','At submission (Outstanding). DR expense (GL Category) / CR 135 AMEX Clearing (Company Card) or 801 Unpaid Expense Claims (Cash).')
row('PE Program','Tag2 = "PE Program" lines reclassed: DR 664 Private Equity Program Expense Receivables (instead of the P&L code).')
row('Totals',f'Company Card -> 135: HKD {card_t:,.2f}.  Cash -> 801: HKD {cash_t:,.2f}.  (of which PE -> 664: HKD {pep_t:,.2f}.)')
row('Journals',f'1 AMEX recognition journal + {len(byrep)} reimbursable journals (one per Report ID, so 801 is tracked per report for settlement).')
row('Not yet booked','AMEX Suspense (855) plug and the AMEX statement Bill require the May AMEX statements + GL opening 855 balance. Payment/Bill step is separate (poster / payment instruction).')
# By card (AMEX)
bc=wb.create_sheet('AMEX by Card'); bc.column_dimensions['A'].width=16; bc.column_dimensions['B'].width=10; bc.column_dimensions['C'].width=16
for j,t in enumerate(['AMEX Card','Lines','Amount (HKD)'],1): sh(bc.cell(1,j,t))
cardsum=defaultdict(lambda:[0,0.0])
for l in lines:
    if l['track']=='CARD': cardsum[l['card']][0]+=1; cardsum[l['card']][1]+=l['amt']
rr=2
for cd in sorted(cardsum):
    bc.cell(rr,1,cd); bc.cell(rr,2,cardsum[cd][0]); bc.cell(rr,3,round(cardsum[cd][1],2)).number_format=HKD; rr+=1
bc.cell(rr,1,'TOTAL -> CR 135').font=F(bold=True); bc.cell(rr,3,f'=SUM(C2:C{rr-1})').number_format=HKD; bc.cell(rr,3).font=F(bold=True)
for cc in (1,2,3): bc.cell(rr,cc).fill=sub
# Open-item schedule (reimbursable by report)
oi=wb.create_sheet('801 Open-Item Schedule'); 
for col,w in zip('ABCDEF',(16,22,14,16,14,12)): oi.column_dimensions[col].width=w
oi['A1']='Unpaid Expense Claims (801) — open items by Expensify Report ID'; oi['A1'].font=F(bold=True,color=NAVY)
oi['A2']='Each report = one 801 balance; cleared when paid (DR 801 / CR DBS x1627) and matched on Report ID.'; oi['A2'].font=F(italic=True,size=9)
for j,t in enumerate(['Report ID','Submitter','Submitted','Report Status','801 Amount (HKD)','Lines'],1): sh(oi.cell(4,j,t))
rr=5; first=rr
for rid in sorted(byrep):
    ls=byrep[rid]; tot=round(sum(x['amt'] for x in ls),2)
    oi.cell(rr,1,rid); oi.cell(rr,2,ls[0]['sub']); oi.cell(rr,3,ls[0]['subdate']); oi.cell(rr,4,ls[0]['rstatus'])
    oi.cell(rr,5,tot).number_format=HKD; oi.cell(rr,6,len(ls)); rr+=1
oi.cell(rr,1,'TOTAL -> CR 801').font=F(bold=True); oi.cell(rr,5,f'=SUM(E{first}:E{rr-1})').number_format=HKD; oi.cell(rr,5).font=F(bold=True)
for cc in range(1,7): oi.cell(rr,cc).fill=sub
# Line detail
ld=wb.create_sheet('Line Detail')
heads=['Date','Merchant','Category','GL code','DR account','Track','Card','Report ID','Amount (HKD)','PE?']
for col,w in zip('ABCDEFGHIJ',(11,26,24,9,10,7,9,15,14,6)): ld.column_dimensions[col].width=w
for j,t in enumerate(heads,1): sh(ld.cell(1,j,t))
rr=2
for l in sorted(lines,key=lambda x:(x['track'],x['rid'],x['date'])):
    vals=[l['date'],l['mer'][:26],l['cat'],l['gl'],l['dr'],l['track'],l['card'],l['rid'],l['amt'],'Y' if l['pe'] else '']
    for j,v in enumerate(vals,1):
        c=ld.cell(rr,j,v); 
        if j==9: c.number_format=HKD
    rr+=1
ld.cell(rr,9,f'=SUM(I2:I{rr-1})').number_format=HKD; ld.cell(rr,9).font=F(bold=True); ld.cell(rr,8,'TOTAL').font=F(bold=True)
# Checklist
ck=wb.create_sheet('Review Checklist'); ck.column_dimensions['A'].width=4; ck.column_dimensions['B'].width=98
ck['B1']=f'Review checklist — {LBL} expenses'; ck['B1'].font=F(bold=True,size=13,color=NAVY); r2=3
for it in [f'{len(lines)} Outstanding lines (1 Draft excluded); all HKD.',
           f'Company Card HKD {card_t:,.2f} -> CR 135; Cash HKD {cash_t:,.2f} -> CR 801; both balance.',
           'P&L code per line = GL Category (col K). PE Program (Tag2) -> DR 664.',
           '23 journals (1 AMEX + 22 per-report reimbursable). 801 tracked per Report ID for settlement.',
           'OPEN: AMEX Suspense (855) plug + AMEX statement Bill — need May AMEX statements + GL opening 855.',
           'OPEN: payment/Bill step (DR 801 or 135 / CR AP) — pending design confirmation.',
           '624 Employee receivables: no personal-charge lines this month.']:
    ck.cell(r2,1,'☐'); ck.cell(r2,2,it).alignment=Alignment(wrap_text=True,vertical='top'); r2+=1
wb.save(f'{OUT}/3CP_EXP_{LBL.replace(" ","")}_Workpaper.xlsx')
print('workpaper saved')
