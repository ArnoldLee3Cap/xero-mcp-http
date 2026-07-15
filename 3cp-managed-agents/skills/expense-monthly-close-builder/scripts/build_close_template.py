"""
build_close_template.py

Generates an Excel month-end close workbook for the 3 Capital Partners expense cycle.

Usage:
    python build_close_template.py --month "March 2026" --output /path/to/output.xlsx --data /path/to/input.json

The input JSON should contain:
{
    "month": "March 2026",
    "amex_statement_total": 100000,
    "amex_payment_date": "2026-03-19",
    "items": [
        {
            "source": "AMEX",
            "date": "2026-03-15",
            "employee": "Alice Wong",
            "merchant": "Restaurant X",
            "amount_hkd": 5000,
            "description": "Client dinner",
            "category": "Reimbursable",
            "pep_fund": null,
            "prepayment_period": null,
            "status": "Submitted",
            "confidence": "High",
            "notes": ""
        },
        ...
    ],
    "prior_month_carryovers": {
        "amex_unsubmitted_suspense_opening": 0,
        "unallocated_suspense_opening": 0,
        "accrued_expenses_opening": 0,
        "items_to_reverse": [
            {"description": "Prior accrual for [X]", "amount": 20000, "reason": "Reclassify to PEP", "new_category": "PEP"}
        ]
    }
}

The script produces a workbook with:
    1. Inputs sheet — raw input data
    2. Classification sheet — line-by-line classification
    3. Reclassifications sheet — corrections from prior months
    4. Summary Journal sheet — the Dr/Cr entry to post
    5. Balance Sheet Check sheet — verification of totals
    6. Suspense Aging sheet — outstanding suspense items by age

If no JSON file is provided, the script generates a blank template with sample data and instructions.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FONT = Font(bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
WARNING_FILL = PatternFill("solid", fgColor="FFC7CE")
CHECK_OK_FILL = PatternFill("solid", fgColor="C6EFCE")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Account names (matches journal_templates.md)
ACCOUNTS = {
    "amex_clearing": "AMEX Clearing",
    "amex_payable": "AMEX Payable",
    "employee_reimbursement_payable": "Employee Reimbursement Payable",
    "accrued_expenses": "Accrued Expenses",
    "receivable_pep": "Receivable from PEP Fund",
    "staff_receivable": "Staff Receivable",
    "prepayment": "Prepayment",
    "amex_unsubmitted_suspense": "AMEX Unsubmitted Suspense",
    "unallocated_suspense": "Unallocated Suspense",
    "expense_reimbursable": "Expense - Reimbursable",
    "expense_subscriptions": "Expense - Subscriptions",
    "cash": "Cash",
}

# Category to account mapping
CATEGORY_TO_ACCOUNT = {
    "Reimbursable": "expense_reimbursable",
    "PEP": "receivable_pep",
    "Non-reimbursable": "staff_receivable",
    "Prepayment": "prepayment",
    "Unsubmitted": "amex_unsubmitted_suspense",
    "Unidentified": "unallocated_suspense",
    "Subscription": "expense_subscriptions",
}

# Payment method to bridging account
PAYMENT_METHOD_TO_BRIDGE = {
    "AMEX": "amex_clearing",
    "Personal Card": "employee_reimbursement_payable",
    "Cash": "employee_reimbursement_payable",
    "Out-of-pocket": "employee_reimbursement_payable",
}


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def style_subheader(cell):
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def style_total(cell):
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER


def autosize_columns(ws, max_width=50):
    # Build a map from column index to max length, skipping merged cells
    col_max_lengths = {}
    for row in ws.iter_rows():
        for cell in row:
            try:
                # Skip merged cells (they don't have column_letter)
                if not hasattr(cell, "column_letter") or cell.column_letter is None:
                    continue
                if cell.value is not None:
                    length = len(str(cell.value))
                    col_letter = cell.column_letter
                    if length > col_max_lengths.get(col_letter, 0):
                        col_max_lengths[col_letter] = length
            except Exception:
                pass
    for col_letter, max_length in col_max_lengths.items():
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 12)


def build_inputs_sheet(wb, data):
    ws = wb.create_sheet("1. Inputs")

    ws["A1"] = f"Month-End Close — {data.get('month', 'TBD')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A3"] = "AMEX Statement Total (HKD)"
    ws["B3"] = data.get("amex_statement_total", 0)
    ws["B3"].number_format = "#,##0"
    style_subheader(ws["A3"])

    ws["A4"] = "AMEX Payment Date"
    ws["B4"] = data.get("amex_payment_date", "TBD")
    style_subheader(ws["A4"])

    ws["A5"] = "Total Items Captured"
    ws["B5"] = len(data.get("items", []))
    style_subheader(ws["A5"])

    ws["A7"] = "Prior Month Carryovers"
    ws["A7"].font = SUBHEADER_FONT
    carry = data.get("prior_month_carryovers", {})
    ws["A8"] = "AMEX Unsubmitted Suspense (opening)"
    ws["B8"] = carry.get("amex_unsubmitted_suspense_opening", 0)
    ws["B8"].number_format = "#,##0"
    ws["A9"] = "Unallocated Suspense (opening)"
    ws["B9"] = carry.get("unallocated_suspense_opening", 0)
    ws["B9"].number_format = "#,##0"
    ws["A10"] = "Accrued Expenses (opening)"
    ws["B10"] = carry.get("accrued_expenses_opening", 0)
    ws["B10"].number_format = "#,##0"

    ws["A12"] = (
        "Instructions: Fill in this sheet with the AMEX statement total, "
        "payment date, and any opening balances from the prior month's close. "
        "All item-level detail goes in the Classification sheet."
    )
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A12:D14")

    autosize_columns(ws)


def build_classification_sheet(wb, data):
    ws = wb.create_sheet("2. Classification")

    headers = [
        "Source",
        "Payment Method",
        "Date",
        "Employee",
        "Merchant",
        "Amount (HKD)",
        "Description",
        "Category",
        "PEP Fund",
        "Prepayment Period",
        "Status",
        "Confidence",
        "GL Account",
        "Credit Account",
        "Notes",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)

    items = data.get("items", [])

    for row_idx, item in enumerate(items, start=2):
        category = item.get("category", "Unidentified")
        gl_account = ACCOUNTS.get(
            CATEGORY_TO_ACCOUNT.get(category, "unallocated_suspense"),
            "Unallocated Suspense",
        )
        payment_method = item.get("payment_method", "AMEX")
        # For AMEX items, the net credit (after AMEX Clearing nets to zero) is Cash
        # For personal card items, the credit is Employee Reimbursement Payable
        bridge_key = PAYMENT_METHOD_TO_BRIDGE.get(payment_method, "amex_clearing")
        if bridge_key == "amex_clearing":
            credit_account = "Cash (via AMEX Clearing)"
        else:
            credit_account = ACCOUNTS["employee_reimbursement_payable"]

        ws.cell(row=row_idx, column=1, value=item.get("source", ""))
        ws.cell(row=row_idx, column=2, value=payment_method)
        ws.cell(row=row_idx, column=3, value=item.get("date", ""))
        ws.cell(row=row_idx, column=4, value=item.get("employee", ""))
        ws.cell(row=row_idx, column=5, value=item.get("merchant", ""))
        amt_cell = ws.cell(row=row_idx, column=6, value=item.get("amount_hkd", 0))
        amt_cell.number_format = "#,##0"
        ws.cell(row=row_idx, column=7, value=item.get("description", ""))
        ws.cell(row=row_idx, column=8, value=category)
        ws.cell(row=row_idx, column=9, value=item.get("pep_fund", "") or "")
        ws.cell(row=row_idx, column=10, value=item.get("prepayment_period", "") or "")
        ws.cell(row=row_idx, column=11, value=item.get("status", ""))
        ws.cell(row=row_idx, column=12, value=item.get("confidence", ""))
        ws.cell(row=row_idx, column=13, value=gl_account)
        ws.cell(row=row_idx, column=14, value=credit_account)
        ws.cell(row=row_idx, column=15, value=item.get("notes", ""))

        # Flag low confidence items
        if item.get("confidence", "") == "Low":
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).fill = WARNING_FILL

    # Add total row
    if items:
        total_row = len(items) + 2
        ws.cell(row=total_row, column=5, value="TOTAL")
        ws.cell(row=total_row, column=5).font = TOTAL_FONT
        total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row - 1})")
        total_cell.font = TOTAL_FONT
        total_cell.number_format = "#,##0"

    autosize_columns(ws)


def build_reclassifications_sheet(wb, data):
    ws = wb.create_sheet("3. Reclassifications")

    ws["A1"] = "Reclassifications from Prior Months"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["Description", "Amount (HKD)", "Reason", "From Account", "To Account", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)

    reversals = data.get("prior_month_carryovers", {}).get("items_to_reverse", [])

    for row_idx, item in enumerate(reversals, start=4):
        ws.cell(row=row_idx, column=1, value=item.get("description", ""))
        amt = ws.cell(row=row_idx, column=2, value=item.get("amount", 0))
        amt.number_format = "#,##0"
        ws.cell(row=row_idx, column=3, value=item.get("reason", ""))

        from_acct = item.get("from_account", "Accrued Expenses")
        to_cat = item.get("new_category", "")
        to_acct = ACCOUNTS.get(CATEGORY_TO_ACCOUNT.get(to_cat, "unallocated_suspense"), to_cat)

        ws.cell(row=row_idx, column=4, value=from_acct)
        ws.cell(row=row_idx, column=5, value=to_acct)
        ws.cell(row=row_idx, column=6, value=item.get("notes", ""))

    if not reversals:
        ws.cell(row=4, column=1, value="No reclassifications this month")
        ws.cell(row=4, column=1).font = Font(italic=True)

    autosize_columns(ws)


def aggregate_for_journal(data):
    """Aggregate classified items into account-level totals for the summary journal.

    Returns a tuple (debit_totals, credit_totals, p_and_l_impact).

    The summary journal combines TWO economic events:
    1. AMEX statement payment (Dr. AMEX Clearing / Cr. Cash)
    2. Allocation of AMEX statement to categories (Dr. various / Cr. AMEX Clearing)

    When combined, AMEX Clearing nets to zero, leaving:
    - Dr. allocation accounts
    - Cr. Cash (for AMEX items)
    - Cr. Employee Reimbursement Payable (for personal card items)
    """
    items = data.get("items", [])
    debit_totals = {acct: 0 for acct in ACCOUNTS.values()}
    credit_totals = {acct: 0 for acct in ACCOUNTS.values()}
    p_and_l_impact = 0
    amex_allocated = 0  # Sum of AMEX-backed items
    personal_allocated = 0  # Sum of personal card / cash items

    for item in items:
        category = item.get("category", "Unidentified")
        amount = item.get("amount_hkd", 0)
        payment_method = item.get("payment_method", "AMEX")
        acct_key = CATEGORY_TO_ACCOUNT.get(category, "unallocated_suspense")
        acct_name = ACCOUNTS.get(acct_key, "Unallocated Suspense")

        # Debit the appropriate expense / asset account
        debit_totals[acct_name] += amount

        # Track totals by payment method for credit side
        if PAYMENT_METHOD_TO_BRIDGE.get(payment_method, "amex_clearing") == "amex_clearing":
            amex_allocated += amount
        else:
            personal_allocated += amount

        # P&L impact for Reimbursable and Subscription only
        if category in ("Reimbursable", "Subscription"):
            p_and_l_impact += amount

    # AMEX items: credit side is Cash (because Company paid AMEX from cash)
    # The AMEX Clearing account is intentionally omitted from the summary journal
    # because it nets to zero (Dr from payment, Cr from allocation).
    # We show the net cash outflow directly.
    if amex_allocated > 0:
        credit_totals[ACCOUNTS["cash"]] = amex_allocated

    # Personal card items: credit side is Employee Reimbursement Payable
    # (Cash flows later when employee is reimbursed)
    if personal_allocated > 0:
        credit_totals[ACCOUNTS["employee_reimbursement_payable"]] = personal_allocated

    return debit_totals, credit_totals, p_and_l_impact


def _posting_currency_and_amount(item):
    """Card (AMEX) items post in HKD at the settled amount; personal/cash items
    post in their original currency. Returns (currency, amount)."""
    pm = item.get("payment_method", "AMEX")
    if PAYMENT_METHOD_TO_BRIDGE.get(pm, "amex_clearing") == "amex_clearing":
        return "HKD", item.get("amount_hkd", 0)
    ccy = item.get("currency") or item.get("original_currency") or "HKD"
    amt = item.get("original_amount")
    if amt is None:
        amt = item.get("amount_hkd", 0) if ccy == "HKD" else item.get("amount", 0)
    return ccy, amt


def emit_contract_payload(data, period=None):
    """Emit the common journal contract consumed by 3cp-xero-journal-poster.

    Option B: AMEX side is a single HKD journal; the reimbursement side splits
    into one journal per currency. Each journal balances within its currency.
    """
    items = data.get("items", [])
    period = period or data.get("period") or ""
    period_end = data.get("period_end") or (period + "-28" if period else "")

    # Buckets: ("AMEX","HKD") and ("REIMB",<ccy>) -> {debits:{acct:amt}, credit_total}
    buckets = {}
    for item in items:
        category = item.get("category", "Unidentified")
        acct_key = CATEGORY_TO_ACCOUNT.get(category, "unallocated_suspense")
        acct_name = ACCOUNTS.get(acct_key, "Unallocated Suspense")
        pm = item.get("payment_method", "AMEX")
        ccy, amt = _posting_currency_and_amount(item)
        if PAYMENT_METHOD_TO_BRIDGE.get(pm, "amex_clearing") == "amex_clearing":
            seg, credit_acct = "AMEX", ACCOUNTS["cash"]
        else:
            seg, credit_acct = "REIMB", ACCOUNTS["employee_reimbursement_payable"]
        key = (seg, ccy)
        b = buckets.setdefault(key, {"debits": {}, "credit_acct": credit_acct, "credit_total": 0})
        b["debits"][acct_name] = b["debits"].get(acct_name, 0) + amt
        b["credit_total"] += amt

    journals = []
    for (seg, ccy), b in sorted(buckets.items()):
        lines = []
        for acct, amt in sorted(b["debits"].items()):
            lines.append({"account": acct, "debit": round(amt, 2), "credit": 0.00})
        lines.append({"account": b["credit_acct"], "debit": 0.00, "credit": round(b["credit_total"], 2)})
        narration = ("AMEX-settled expenses allocated to categories (HKD)"
                     if seg == "AMEX" else
                     f"Reimbursable out-of-pocket expenses, {ccy}")
        journals.append({
            "reference": f"3CP-EXP-{period}-{seg}-{ccy}",
            "status": "DRAFT",
            "date": period_end,
            "currency": ccy,
            "narration": narration,
            "reversal": {"flag": False, "date": None},
            "lines": lines,
        })

    return {
        "contract_version": "1.0",
        "entity": "3 Capital Partners Limited",
        "source_skill": "expense-monthly-close-builder",
        "period": period,
        "period_end": period_end,
        "target": {"system": "Xero", "post_as": "DRAFT", "auto_approve": False},
        "journals": journals,
    }


def build_summary_journal_sheet(wb, data):
    ws = wb.create_sheet("4. Summary Journal")

    ws["A1"] = f"Summary Journal Entry — {data.get('month', 'TBD')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    headers = ["Account", "Debit (HKD)", "Credit (HKD)", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)

    debit_totals, credit_totals, p_and_l_impact = aggregate_for_journal(data)

    row = 4
    total_debits = 0
    total_credits = 0
    reclassification_pl_impact = 0

    # Process debit accounts
    for acct_name, amount in debit_totals.items():
        if amount == 0:
            continue
        ws.cell(row=row, column=1, value=acct_name)
        cell = ws.cell(row=row, column=2, value=amount)
        cell.number_format = "#,##0"
        total_debits += amount
        row += 1

    # Process credit accounts
    for acct_name, amount in credit_totals.items():
        if amount == 0:
            continue
        ws.cell(row=row, column=1, value=acct_name)
        cell = ws.cell(row=row, column=3, value=amount)
        cell.number_format = "#,##0"
        total_credits += amount
        row += 1

    # Add reclassification entries
    reversals = data.get("prior_month_carryovers", {}).get("items_to_reverse", [])
    if reversals:
        row += 1
        ws.cell(row=row, column=1, value="--- Reclassifications ---")
        ws.cell(row=row, column=1).font = SUBHEADER_FONT
        row += 1
        for item in reversals:
            amount = item.get("amount", 0)
            from_acct = item.get("from_account", "Accrued Expenses")
            to_cat = item.get("new_category", "")
            to_acct = ACCOUNTS.get(
                CATEGORY_TO_ACCOUNT.get(to_cat, "unallocated_suspense"), to_cat
            )

            # Dr to new account
            ws.cell(row=row, column=1, value=to_acct)
            cell = ws.cell(row=row, column=2, value=amount)
            cell.number_format = "#,##0"
            ws.cell(row=row, column=4, value=f"Reclassify: {item.get('reason', '')}")
            total_debits += amount
            row += 1

            # Cr from old account
            ws.cell(row=row, column=1, value=from_acct)
            cell = ws.cell(row=row, column=3, value=amount)
            cell.number_format = "#,##0"
            total_credits += amount
            row += 1

            # Track P&L impact of reclassification
            if from_acct in (ACCOUNTS["accrued_expenses"], ACCOUNTS["expense_reimbursable"], ACCOUNTS["expense_subscriptions"]):
                if to_cat in ("PEP", "Non-reimbursable", "Unidentified", "Unsubmitted"):
                    reclassification_pl_impact -= amount

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    style_total(ws.cell(row=row, column=1))
    dr_cell = ws.cell(row=row, column=2, value=total_debits)
    dr_cell.number_format = "#,##0"
    style_total(dr_cell)
    cr_cell = ws.cell(row=row, column=3, value=total_credits)
    cr_cell.number_format = "#,##0"
    style_total(cr_cell)

    # Balance check
    row += 2
    ws.cell(row=row, column=1, value="Balance Check (Dr - Cr)")
    diff = total_debits - total_credits
    check_cell = ws.cell(row=row, column=2, value=diff)
    check_cell.number_format = "#,##0"
    if diff == 0:
        check_cell.fill = CHECK_OK_FILL
        ws.cell(row=row, column=4, value="✓ Balanced")
    else:
        check_cell.fill = WARNING_FILL
        ws.cell(row=row, column=4, value="✗ NOT BALANCED — review")

    # P&L impact - new items
    row += 1
    ws.cell(row=row, column=1, value="P&L Impact (new items)")
    pl_cell = ws.cell(row=row, column=2, value=p_and_l_impact)
    pl_cell.number_format = "#,##0"

    # P&L impact - reclassifications
    if reclassification_pl_impact != 0:
        row += 1
        ws.cell(row=row, column=1, value="P&L Impact (reclassifications)")
        rec_cell = ws.cell(row=row, column=2, value=reclassification_pl_impact)
        rec_cell.number_format = "#,##0;[Red]-#,##0"

        # Net P&L impact
        row += 1
        ws.cell(row=row, column=1, value="Net P&L Impact this Month")
        net_cell = ws.cell(row=row, column=2, value=p_and_l_impact + reclassification_pl_impact)
        net_cell.number_format = "#,##0;[Red]-#,##0"
        style_total(net_cell)
        style_total(ws.cell(row=row, column=1))
    else:
        style_total(pl_cell)
        style_total(ws.cell(row=row, column=1))

    autosize_columns(ws)


def build_balance_check_sheet(wb, data):
    ws = wb.create_sheet("5. Balance Sheet Check")

    ws["A1"] = "Balance Sheet Reconciliation"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    headers = ["Account", "Movement This Month (HKD)", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)

    debit_totals, credit_totals, p_and_l_impact = aggregate_for_journal(data)

    row = 4
    # Cash out (from the aggregated AMEX credit)
    cash_credit = credit_totals.get(ACCOUNTS["cash"], 0)
    if cash_credit > 0:
        ws.cell(row=row, column=1, value="Cash")
        cell = ws.cell(row=row, column=2, value=-cash_credit)
        cell.number_format = "#,##0"
        ws.cell(row=row, column=3, value="Paid to AMEX (AMEX Clearing nets to zero)")
        row += 1

    # Receivables and Suspense (from debit totals)
    for acct in [
        "receivable_pep",
        "staff_receivable",
        "amex_unsubmitted_suspense",
        "unallocated_suspense",
    ]:
        acct_name = ACCOUNTS[acct]
        amount = debit_totals.get(acct_name, 0)
        if amount != 0:
            ws.cell(row=row, column=1, value=acct_name)
            cell = ws.cell(row=row, column=2, value=amount)
            cell.number_format = "#,##0"
            row += 1

    # Employee Reimbursement Payable (liability - shown as negative)
    erp_amount = credit_totals.get(ACCOUNTS["employee_reimbursement_payable"], 0)
    if erp_amount > 0:
        ws.cell(row=row, column=1, value="Employee Reimbursement Payable")
        cell = ws.cell(row=row, column=2, value=-erp_amount)
        cell.number_format = "#,##0"
        ws.cell(row=row, column=3, value="Liability — cash leaves when employee paid")
        row += 1

    # P&L impact (Retained Earnings)
    ws.cell(row=row, column=1, value="P&L Impact (Retained Earnings)")
    cell = ws.cell(row=row, column=2, value=-p_and_l_impact)
    cell.number_format = "#,##0"
    ws.cell(row=row, column=3, value="Expense reduces equity")
    row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="NET MOVEMENT")
    style_total(ws.cell(row=row, column=1))

    sum_formula = f"=SUM(B4:B{row - 2})"
    cell = ws.cell(row=row, column=2, value=sum_formula)
    cell.number_format = "#,##0"
    style_total(cell)

    row += 1
    ws.cell(row=row, column=1, value="Check (should be 0)")
    ws.cell(row=row, column=3, value="If non-zero, an item has not been correctly classified")

    autosize_columns(ws)


def build_suspense_aging_sheet(wb, data):
    ws = wb.create_sheet("6. Suspense Aging")

    ws["A1"] = "Suspense Account Aging"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    headers = ["Item", "Amount (HKD)", "Original Date", "Age (Days)", "Action"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        style_header(cell)

    items = data.get("items", [])
    today = datetime.now().date()

    row = 4
    for item in items:
        if item.get("category") in ("Unsubmitted", "Unidentified"):
            ws.cell(row=row, column=1, value=item.get("description", item.get("merchant", "")))
            amt = ws.cell(row=row, column=2, value=item.get("amount_hkd", 0))
            amt.number_format = "#,##0"
            ws.cell(row=row, column=3, value=item.get("date", ""))

            try:
                item_date = datetime.strptime(item.get("date", ""), "%Y-%m-%d").date()
                age = (today - item_date).days
                ws.cell(row=row, column=4, value=age)

                if age > 60:
                    action = "ESCALATE to management"
                    ws.cell(row=row, column=5).fill = WARNING_FILL
                elif age > 30:
                    action = "Chase employee"
                    ws.cell(row=row, column=5).fill = WARNING_FILL
                else:
                    action = "Monitor"
                ws.cell(row=row, column=5, value=action)
            except (ValueError, TypeError):
                ws.cell(row=row, column=4, value="—")
                ws.cell(row=row, column=5, value="Check date format")

            row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="No suspense items this month")
        ws.cell(row=4, column=1).font = Font(italic=True)

    autosize_columns(ws)


def build_workbook(data, output_path):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_inputs_sheet(wb, data)
    build_classification_sheet(wb, data)
    build_reclassifications_sheet(wb, data)
    build_summary_journal_sheet(wb, data)
    build_balance_check_sheet(wb, data)
    build_suspense_aging_sheet(wb, data)

    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")


def get_sample_data():
    """Returns a sample dataset matching Example 1 from worked_examples.md, with added personal card items."""
    return {
        "month": "March 2026",
        "amex_statement_total": 100000,
        "amex_payment_date": "2026-03-19",
        "items": [
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-03-05",
                "employee": "Alice Wong",
                "merchant": "Hotel Tokyo",
                "amount_hkd": 50000,
                "description": "PEP fund due diligence trip",
                "category": "PEP",
                "pep_fund": "PEP Fund I",
                "prepayment_period": None,
                "status": "Submitted",
                "confidence": "High",
                "notes": "Tagged PEP at submission",
            },
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-03-10",
                "employee": "Brian Lee",
                "merchant": "Client Restaurant",
                "amount_hkd": 20000,
                "description": "Client entertainment dinner",
                "category": "Reimbursable",
                "pep_fund": None,
                "prepayment_period": None,
                "status": "Submitted",
                "confidence": "Medium",
                "notes": "May need PEP review",
            },
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-03-15",
                "employee": "Carol Chan",
                "merchant": "Unknown vendor",
                "amount_hkd": 10000,
                "description": "Charge on AMEX, no Expensify submission",
                "category": "Unsubmitted",
                "pep_fund": None,
                "prepayment_period": None,
                "status": "Pending",
                "confidence": "Low",
                "notes": "Chase Carol",
            },
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-01-05",
                "employee": "David Wu",
                "merchant": "Microsoft 365",
                "amount_hkd": 10000,
                "description": "Q1 subscription Jan-Mar",
                "category": "Subscription",
                "pep_fund": None,
                "prepayment_period": "2026-01-01 to 2026-03-31",
                "status": "Approved",
                "confidence": "High",
                "notes": "Amortise across Q1",
            },
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-03-20",
                "employee": "Emily Cheung",
                "merchant": "Personal store",
                "amount_hkd": 8000,
                "description": "Personal shopping — non-reimbursable",
                "category": "Non-reimbursable",
                "pep_fund": None,
                "prepayment_period": None,
                "status": "Rejected",
                "confidence": "High",
                "notes": "Employee to repay",
            },
            {
                "source": "AMEX",
                "payment_method": "AMEX",
                "date": "2026-03-25",
                "employee": None,
                "merchant": "Unknown",
                "amount_hkd": 2000,
                "description": "Unidentified AMEX charge",
                "category": "Unidentified",
                "pep_fund": None,
                "prepayment_period": None,
                "status": "Pending",
                "confidence": "Low",
                "notes": "Investigate",
            },
            {
                "source": "Expensify",
                "payment_method": "Personal Card",
                "date": "2026-03-18",
                "employee": "Frank Ng",
                "merchant": "Taxi",
                "amount_hkd": 3000,
                "description": "Client meeting transport",
                "category": "Reimbursable",
                "pep_fund": None,
                "prepayment_period": None,
                "status": "Approved",
                "confidence": "High",
                "notes": "Personal card — to be reimbursed",
            },
            {
                "source": "Expensify",
                "payment_method": "Personal Card",
                "date": "2026-03-22",
                "employee": "Grace So",
                "merchant": "PEP Conference Fee",
                "amount_hkd": 5000,
                "description": "PEP fund LP conference registration",
                "category": "PEP",
                "pep_fund": "PEP Fund II",
                "prepayment_period": None,
                "status": "Approved",
                "confidence": "High",
                "notes": "Personal card — reclassify to PEP",
            },
        ],
        "prior_month_carryovers": {
            "amex_unsubmitted_suspense_opening": 0,
            "unallocated_suspense_opening": 0,
            "accrued_expenses_opening": 0,
            "items_to_reverse": [],
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Build month-end close workbook")
    parser.add_argument("--month", default="Current Month", help="Month label for the workbook")
    parser.add_argument(
        "--output", required=True, help="Output xlsx file path"
    )
    parser.add_argument(
        "--data",
        default=None,
        help="Path to input JSON file. If omitted, sample data is used.",
    )
    parser.add_argument("--period", default=None, help="Accounting period YYYY-MM (for the contract payload references)")
    parser.add_argument("--emit-payload", default=None, help="Also write the contract JSON payload (for 3cp-xero-journal-poster) to this path")
    args = parser.parse_args()

    if args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "month" not in data:
            data["month"] = args.month
    else:
        data = get_sample_data()
        data["month"] = args.month
        print("No input data provided — generating with sample data (Example 1 from skill).")

    build_workbook(data, args.output)

    if args.emit_payload:
        payload = emit_contract_payload(data, period=args.period)
        with open(args.emit_payload, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)
        print(f"Contract payload written to {args.emit_payload} ({len(payload['journals'])} journal(s))")


if __name__ == "__main__":
    main()
